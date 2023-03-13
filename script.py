import openpyxl
import pandas as pd
from os import getcwd, startfile, path, listdir
import fill_client_dict, connect
from openpyxl.styles import PatternFill, Font, Alignment, Side, Border
from openpyxl.formatting.rule import ColorScaleRule
import send_email


# Normalized list will have the data we need from the csvs organized as we want it to be able to manipulate it.
normalized_list = []

#fill client dict will get a dictionary with all the Permanents with their clients, origins, and destinations.
perm_dict = fill_client_dict.main()

current_directory = getcwd()


dirname = r'{current_directory}\\GeneratedCSVs'.format(current_directory = current_directory)

print(dirname)

weeks = connect.main(current_directory)

file_name = 'Report from {date}.xlsx'.format(date = weeks[3][:6] + ' - ' + weeks[0][-6:])

list_of_files = sorted( filter( lambda x: path.isfile(path.join(dirname, x)),
                        listdir(dirname) ) )

# Loop through the files corresponding to the last 4 weeks.
for i in range(0,4):

    ####################### USING PANDAS #####################

    # We set our dataframe to be equal to the csv the user selected.
    df = pd.read_csv('{directory}\\{file}'.format(directory = dirname, file = list_of_files[i]))

    # At the moment the first row is always row number 6
    first_row = 6
    # We get the last row by looking for the index of the first row that has the text 'Permanent Circuits Availability' on the second column
    last_row = df[df.iloc[:, 1] == 'Permanent Circuits Availability'].index[0]

    pd.options.display.max_rows = 100
    pd.options.display.max_columns = 10

    # We add a boolean to let us know if we are still on the first page or not since the column position changes on pages that are not the first
    first_page = True

    # We get the info we want from the csv and pass it into a list, then we put that list into normalized list as a new item
    for j in range(first_row,last_row):
        if df.iloc[j,0] == 'Generated on:' or df.iloc[j,1] in ['Permanent Circuits Outages', 'Resolution']:
            first_page = False
            continue
        if first_page:
            tmp_lst = [df.iloc[j, 1], df.iloc[j, 3], df.iloc[j, 6], df.iloc[j, 5], df.iloc[j,20], str(i+1)]
        else:
            tmp_lst = [df.iloc[j, 1], df.iloc[j, 3], df.iloc[j, 5], df.iloc[j, 4], df.iloc[j,16], str(i+1)]
            # Check that there ir a permanent on the line.
            if pd.isna(tmp_lst[1]):
                normalized_list[-1][4] += tmp_lst[4]
                continue
        #print(tmp_lst)
        normalized_list.append(tmp_lst)

# Headers of our dataframe
header = ['Resolution', 'Permanent', 'Issue', 'Jira', 'Root Cause', 'Week']

df2 = pd.DataFrame(normalized_list).fillna(value="---> NO INFO <---")
df2.columns = header

# print(df2)

while True:
    try:
        df2.to_excel('source information.xlsx',header=True, index= False)
        break
    except PermissionError:
        input('Close the file before attempting to create a new one.\nPress Enter once you have closed the file to attempt to create it again.')
        continue

wb = openpyxl.load_workbook('source information.xlsx')
ws = wb['Sheet1']
ws.auto_filter.ref = 'A1:F1'
ws.column_dimensions['A'].width = 18.29
ws.column_dimensions['B'].width = 38.57
ws.column_dimensions['C'].width = 20.29
ws.column_dimensions['D'].width = 11.14
ws.column_dimensions['E'].width = 16.57
ws.column_dimensions['F'].width = 10.14
while True:
    try:
        wb.save('source information.xlsx')
        break
    except OSError:
        input('Close the file before attempting to create a new one.\nPress Enter once you have closed the file to attempt to create it again.')
        continue
input('\nA file with the source information has been created. \nPlease take a moment to check the source info file since this is the information that will be used to generate the report.\nPress Enter to open the file')
try:
    file = current_directory + "\\{file}".format(file = 'source information.xlsx')
    print ("Attempting to open file")
    startfile(file)
except FileNotFoundError:
    print("Unable to open file, try opening it normally.")


input('\nPress Enter if you want to continue')


# Create a list with only one instance of each permanent
permanents = []

for line in normalized_list:
    # We check if the permanent is already on the permanent list, if so we continue if not we add it
    if line[1] in permanents:
        continue
    # If not we add it to the list.
    permanents.append(line[1])

issue_list = []

for permanent in permanents:
    # Initialize our counters for each issue that we want to keep track of.
    equipment1 = 0
    link_degradation1 = 0
    link_outage1 = 0
    equipment2 = 0
    link_degradation2 = 0
    link_outage2 = 0
    equipment3 = 0
    link_degradation3 = 0
    link_outage3 = 0
    equipment4 = 0
    link_degradation4 = 0
    link_outage4 = 0
    total_equipment = 0
    total_degradation = 0
    total_outage = 0
    total = 0
    root_cause_summary = ''
    for line in normalized_list:
        if line[0] == 'Duplicate':
            continue
        # We extract what was the issue for each line and in which week happened.
        issue = line[2]
        root_cause = line[4]
        if pd.isna(root_cause):
            root_cause = ''
        root_cause = root_cause.strip()
        week = line[5]
        # We check that the permanent we are checking is the one being affected in this line
        if permanent == line[1] and week == '1':# if the permanent we are checking matches the permanent in this line and corresponds to week 1
            match issue:
                # We check the issue if it's any of the ones we care about we add into the particular counter.
                case 'Equipment Problem': equipment1 += 1; total += 1; total_equipment +=1 #; root_cause_summary += root_cause + '\n\n'
                case 'Link Degradation': link_degradation1 += 1; total += 1; total_degradation +=1; #root_cause_summary += root_cause + '\n\n'
                case 'Link Outage': link_outage1 += 1; total += 1; total_outage += 1; #root_cause_summary += root_cause + '\n\n'
                case _: continue
        elif permanent == line[1] and week == '2':# if the permanent we are checking matches the permanent in this line and corresponds to week 2
            match issue:
                # We check the issue if it's any of the ones we care about we add into the particular counter.
                case 'Equipment Problem': equipment2 += 1; total += 1; total_equipment +=1; #root_cause_summary += root_cause + '\n\n'
                case 'Link Degradation': link_degradation2 += 1; total += 1; total_degradation +=1; #root_cause_summary += root_cause + '\n\n'
                case 'Link Outage': link_outage2 += 1; total += 1; total_outage += 1; #root_cause_summary += root_cause + '\n\n'
                case _: continue
        elif permanent == line[1] and week == '3':# if the permanent we are checking matches the permanent in this line and corresponds to week 3
            match issue:
                # We check the issue if it's any of the ones we care about we add into the particular counter.
                case 'Equipment Problem': equipment3 += 1; total += 1; total_equipment +=1; #root_cause_summary += root_cause + '\n\n'
                case 'Link Degradation': link_degradation3 += 1; total += 1; total_degradation +=1; #root_cause_summary += root_cause + '\n\n'
                case 'Link Outage': link_outage3 += 1; total += 1; total_outage += 1; #root_cause_summary += root_cause + '\n\n'
                case _: continue
        elif permanent == line[1] and week == '4':# if the permanent we are checking matches the permanent in this line and corresponds to week 3
            match issue:
                # We check the issue if it's any of the ones we care about we add into the particular counter.
                case 'Equipment Problem': equipment4 += 1; total += 1; total_equipment +=1; root_cause_summary += r'•   ' + root_cause + '\n'
                case 'Link Degradation': link_degradation4 += 1 ; total += 1; total_degradation +=1; root_cause_summary += r'•   ' + root_cause + '\n'
                case 'Link Outage': link_outage4 += 1; total += 1; total_outage += 1; root_cause_summary += r'•   ' + root_cause + '\n'
                case _: continue
        else:
            continue
    try:
        customer = perm_dict[permanent]['Customer']
    except KeyError:
        for key, value in perm_dict.items():
            if key.startswith(permanent):
                customer = value['Customer']
                break
            else: customer = 'Customer not found, add manually'
    try:
        origin = perm_dict[permanent]['Origin']
    except KeyError:
        for key, value in perm_dict.items():
            if key.startswith(permanent):
                origin = value['Origin']
                break
            else: origin = 'Origin not found, add manually'
    try:
        destination = perm_dict[permanent]['Destinations'].rstrip(' // ').replace(' //',',\n')
    except KeyError:
        for key, value in perm_dict.items():
            if key.startswith(permanent):
                destination = value['Destinations'].rstrip(' // ').replace(' //',',\n')
                break
            else: destination = 'Destination not found, add manually'
    # We pass the information to our issue list with the final data.
    if equipment4 == 0 and link_degradation4 == 0 and link_outage4 == 0 and total >=3: root_cause_summary = 'No issues last week'
    if total<3: root_cause_summary=''
    if total == 0: continue
    root_cause_summary = root_cause_summary.strip()
    tmp_lst = [permanent, customer, origin, destination, equipment1, link_degradation1, link_outage1, equipment2, link_degradation2, link_outage2, equipment3, link_degradation3, link_outage3, equipment4, link_degradation4, link_outage4, total_equipment, total_degradation, total_outage, total, root_cause_summary]
    issue_list.append(tmp_lst)

# Generate dataframe from list and write to xlsx.

# Creating a header list to be the headers on the final excel file.
header = ['Affected Customer Circuit','Client', 'Origin', 'Destination', 'Equipment Problem','Link Degradation','Link Outage','Equipment Problem','Link Degradation','Link Outage','Equipment Problem','Link Degradation','Link Outage', 'Equipment Problem','Link Degradation','Link Outage','Equipment Problem','Link Degradation','Link Outage','Total', 'Actions Taken']
# Passing the issue_list as the dataframe to be written into the excel file.
df = pd.DataFrame(issue_list)

# Adding a header to our dataframe (df).
df.columns = header
df = df.sort_values(by='Total', ascending=False)
# print(df)
# Creating our excel file.
while True:
    try:
        df.to_excel(file_name,header=True, index= False)
        break
    except PermissionError:
        input('Close the file before attempting to create a new one.\nPress Enter once you have closed the file to attempt to create it again.')
        continue

# Styling the final document

# First open the document with the information.
wb = openpyxl.load_workbook(file_name)
final_df = pd.read_excel(file_name)
number_of_rows = df.shape[0]
ws = wb['Sheet1']
ws.page_setup.fitToHeight = 1
ws.page_setup.fitToWidth = 1
# Setup the width of the columns.
ws.column_dimensions['A'].width = 23.57
ws.column_dimensions['B'].width = 33.14
ws.column_dimensions['C'].width = 29.86
ws.column_dimensions['D'].width = 44.14
ws.column_dimensions['U'].width = 78.43
ws.column_dimensions['V'].width = 78.43
# Fill the headers with the selected color.
fill_cell = PatternFill(patternType='solid', fgColor='DCE6F1')
# Define the font and borders
text_color = Font(color='000000', bold=True)
medium = Side(border_style="medium", color="000000")
thin = Side(border_style="thin", color="000000")
# Adding comments columns 
ws['V1'] = 'Comments'
ws['V1'].border = Border(top=thin, left=thin, right= thin, bottom=thin) 
# Set up filters
ws.auto_filter.ref = 'A3:V3'
column_name = ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U','V']
vertical_cells = ['E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T']
# Apply styles to the headers.
for cell in column_name:
    ws[cell+'1'].fill = fill_cell
    ws[cell+'1'].font = text_color
    ws[cell+'1'].alignment = Alignment(horizontal='center', vertical='center')
    if cell in vertical_cells:
        ws.column_dimensions[cell].width = 5
        ws[cell+'1'].alignment = Alignment(text_rotation=90, horizontal='center', vertical='center')
    for i in range (1, number_of_rows + 2):
        ws[cell+str(i)].border = Border(top=thin, left=thin, right= thin, bottom=thin)

# Add rows to write new information above of the table.
ws.insert_rows(1,2)

center = Alignment(horizontal="center", vertical="center")
medium_border = Border(top=medium, left=medium, right= medium, bottom=medium) 
calibri_10_bold = Font('Calibri', 10, color='000000', bold= True)

ws['A1'] = 'Fever Client {weeks}'.format(weeks = weeks[3][:6] + ' - ' + weeks[0][-6:])
ws['A1'].font = calibri_10_bold
ws['A1'].alignment = center
ws['A1'].border = medium_border
ws.merge_cells('A1:A2')

ws['B1'].alignment = center
ws['B1'].border = medium_border
ws.merge_cells('B1:D2')

ws['E1'] = 'Week'
ws['E1'].font = calibri_10_bold
ws['E1'].alignment = center
ws['E1'].border = medium_border
ws.merge_cells('E1:P1')

ws['E2'] = weeks[3]
ws['E2'].font = calibri_10_bold
ws['E2'].alignment = center
ws['E2'].border = medium_border
ws.merge_cells('E2:G2')

ws['H2'] = weeks[2]
ws['H2'].font = calibri_10_bold
ws['H2'].alignment = center
ws['H2'].border = medium_border
ws.merge_cells('H2:J2')

ws['K2'] = weeks[1]
ws['K2'].font = calibri_10_bold
ws['K2'].alignment = center
ws['K2'].border = medium_border
ws.merge_cells('K2:M2')

ws['N2'] = weeks[0]
ws['N2'].font = calibri_10_bold
ws['N2'].alignment = center
ws['N2'].border = medium_border
ws.merge_cells('N2:P2')

ws['Q1'] = 'Total'
ws['Q1'].font = calibri_10_bold
ws['Q1'].alignment = center
ws['Q1'].border = medium_border
ws.merge_cells('Q1:T2')

# Rule to create the color scale gradiant for the total columns
rule = ColorScaleRule(start_type='min', start_color='63BE7B',
                       mid_type='percentile', mid_value=50, mid_color='FFEB84',
                       end_type='max', end_color='F8696B')

for cell in ws["T:T"]:
    conditional_cells = cell.row
    cell.font = text_color    

ws.auto_filter.add_sort_condition("T:T",descending=True)
ws.conditional_formatting.add("Q4:T{conditional}".format(conditional = conditional_cells), rule)

#Save the styled file.
while True:
    try:
        wb.save(file_name)
        break
    except OSError:
        input('Close the file before attempting to create a new one.\nPress Enter once you have closed the file to attempt to create it again.')
        continue
        
# We tell the user where the file is located.
print("File generated on", current_directory)

input("Press Enter to open the file and close this window.")

# We try to open the file so the user can see it right away.
try:
    file = current_directory + "\\{file}".format(file = file_name)
    print ("Attempting to open file")
    startfile(file)
except FileNotFoundError:
    print("Unable to open file, try opening it normally.")

input('Press Enter to send the email once you finish editing the file')

send_email.main(file,file_name)